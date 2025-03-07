import signal

import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.workbook import Workbook

from constants import *

thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

class BattleFileExcel():

    def __init__(self,FILENAME,player1,player2):

        try:
            self.file = load_workbook(FILENAME)
        except FileNotFoundError:
            self.file = Workbook()
            self.file.save(FILENAME)
            self.file.create_sheet(title='Resultados')
            self.file.create_sheet(title='Estadisticas')

        self.datos = dict()
        self.player1data=dict()
        self.player2data = dict()

        self.player1data['R']=0
        self.player1data['P'] = 0
        self.player1data['T'] = 0

        self.player2data['R'] = 0
        self.player2data['P'] = 0
        self.player2data['T'] = 0

        self.player1=player1
        self.FILENAME=FILENAME
        self.player2=player2

        self.datos[player1]=[]
        self.datos[player2]=[]
        self.datos['Ganador']=[]


        signal.signal(signal.SIGINT, self.handleExit)  # Captura Ctrl+C
        signal.signal(signal.SIGTERM, self.handleExit)  # Captura kill PID

    def writeRound(self,c1,c2,winner):
        self.datos[self.player1].append(c1)
        self.datos[self.player2].append(c2)
        self.datos['Ganador'].append(winner)

        self.player1data[c1]+=1
        self.player2data[c1] += 1


    def createExcelHistory(self):
        ws = self.file['Resultados']
        ws.cell(row=1, column=1, value=self.player1).fill= PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        ws.cell(row=1, column=2, value=self.player2).fill= PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        ws.cell(row=1, column=3, value="Ganador").fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        ws.cell(row=1, column=1).border = thin_border
        ws.cell(row=1, column=2).border = thin_border
        ws.cell(row=1, column=3).border = thin_border

    def createScoreboardTable(self):
        ws = self.file['Resultados']
        ws.cell(row=SCOREBOARD_ROW,column=SCOREBOARD_COLUMN_INIT,value=self.player1).border = thin_border
        ws.cell(row=SCOREBOARD_ROW, column=SCOREBOARD_COLUMN_INIT+1, value=self.player2).border = thin_border
        ws.cell(row=SCOREBOARD_ROW+1, column=SCOREBOARD_COLUMN_INIT, value=0).border = thin_border
        ws.cell(row=SCOREBOARD_ROW+1, column=SCOREBOARD_COLUMN_INIT+1, value=0).border = thin_border


    def updateScoreboardTable(self,score1,score2):
        ws = self.file['Resultados']
        ws.cell(row=SCOREBOARD_ROW + 1, column=SCOREBOARD_COLUMN_INIT, value=score1)
        ws.cell(row=SCOREBOARD_ROW + 1, column=SCOREBOARD_COLUMN_INIT+1, value=score2)


    def updateExcelHistory(self,round,move1,move2,winner):
        ws = self.file['Resultados']
        ws.cell(row=round + 1, column=PLAYER1_HISTORY_COLUMN, value=move1).border = thin_border
        ws.cell(row=round + 1, column=PLAYER2_HISTORY_COLUMN, value=move2).border = thin_border
        ws.cell(row=round + 1, column=WINNER_HISTORY_COLUMN, value=winner).border = thin_border

        if winner == self.player1:
            ws.cell(row=round + 1, column=WINNER_HISTORY_COLUMN).fill=PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif winner == self.player2:
            ws.cell(row=round + 1, column=WINNER_HISTORY_COLUMN).fill = PatternFill(start_color="FFFF99", end_color="FFFF99",fill_type="solid")
        else:
            ws.cell(row=round + 1, column=WINNER_HISTORY_COLUMN).fill = PatternFill(start_color="D3D3D3",end_color="D3D3D3",fill_type="solid")

    def saveExcel(self):
        self.file.save(FILENAME)
        self.file.close()

    def makeGraphics(self):
        ws = self.file['Estadisticas']

        ws.append(["Move", "Valor"])
        for move, value in self.player1data.items():
            ws.append([move, value])

        chart = BarChart()
        data = Reference(ws, min_col=2, min_row=1, max_col=2,
                         max_row=len(self.player1data) + 1)
        categories = Reference(ws, min_col=1, min_row=2,
                               max_row=len(self.player1data) + 1)

        chart.y_axis.majorGridlines = None
        chart.y_axis.majorTickMark = 'in'
        chart.y_axis.minorTickMark = 'none'

        chart.y_axis.min = 0  # Valor mínimo del eje Y
        chart.y_axis.max = ROUNDS  # Valor máximo del eje Y
        chart.y_axis.majorUnit = ROUNDS // 10  # La frecuencia de las marcas cada 100 unidades

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws.add_chart(chart, "E5")
    def handleExit(self,signum,frame):
        self.saveExcel()
        exit(0)
